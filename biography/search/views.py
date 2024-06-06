from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_page
from django.http import HttpResponse
from django.db.models import Q

import openpyxl
from io import BytesIO

from bio.models import Bio
from settings.models import Settings

from search.query import text

# Create your views here.

# @cache_page(0)
def search(request):

    query = request.GET.get('query')

    if query != '' and query != None:
        bios = Bio.objects.filter(text(query))
    else:
        return redirect('homepage')

    response = {
        'query': query,
        'bios': bios,
    }

    if not request.user.is_anonymous:
        response['settings'] = get_object_or_404(Settings, user=request.user)

    return render(request, 'search.html', response)

def group(request):

    if request.LANGUAGE_CODE == 'en':
        college = request.GET.get('college')

        if college == 'All':
            bios = Bio.objects.all()
        else:
            bios = Bio.objects.filter(Q(college=college))

        response = {
            'college': college,
            'bios': bios,
        }

    else:
        ar_college = request.GET.get('ar_college')
        ar_academic_title = request.GET.get('ar_academic_title')

        if ar_college == 'All':
            bios = Bio.objects.all()
        else:
            bios = Bio.objects.filter(Q(ar_college=ar_college))

        response = {
            'ar_college': ar_college,
            'bios': bios,
        }

    if not request.user.is_anonymous:
        response['settings'] = get_object_or_404(Settings, user=request.user)

    return render(request, 'search.html', response)

@login_required
def excel(request):

    if not request.user.is_superuser:
        return redirect('homepage')

    if request.LANGUAGE_CODE == 'en':
        college = request.GET.get('college')
    else:
        college = request.GET.get('ar_college')

    if college == 'All':
        bios = Bio.objects.all()
    else:
        bios = Bio.objects.filter(Q(college=college))

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=2).value =  'ar_full_name'
    sheet.cell(row=1, column=3).value =  'full_name'
    sheet.cell(row=1, column=4).value =  'ar_nationality'
    sheet.cell(row=1, column=5).value =  'ar_college'
    sheet.cell(row=1, column=6).value =  'ar_department'
    sheet.cell(row=1, column=7).value = 'ar_mother_lang'
    sheet.cell(row=1, column=8).value = 'ar_other_langs'
    sheet.cell(row=1, column=9).value = 'ar_state'
    sheet.cell(row=1, column=10).value = 'ar_district'
    sheet.cell(row=1, column=11).value = 'surname'
    sheet.cell(row=1, column=12).value = 'birthday'
    sheet.cell(row=1, column=13).value = 'nationality'
    sheet.cell(row=1, column=14).value = 'gender'
    sheet.cell(row=1, column=15).value = 'degree'
    sheet.cell(row=1, column=16).value = 'college'
    sheet.cell(row=1, column=17).value = 'department'
    sheet.cell(row=1, column=18).value = 'mother_lang'
    sheet.cell(row=1, column=19).value = 'other_langs'
    sheet.cell(row=1, column=20).value = 'email'
    sheet.cell(row=1, column=21).value = 'create_date'
    sheet.cell(row=1, column=22).value = 'update_date'
        
    for i, bio in enumerate(bios, 2):

        sheet.cell(row=i, column=2).value = bio.ar_full_name
        sheet.cell(row=i, column=3).value = bio.full_name
        sheet.cell(row=i, column=4).value = bio.ar_nationality
        sheet.cell(row=i, column=5).value = bio.ar_college
        sheet.cell(row=i, column=6).value = bio.ar_department
        sheet.cell(row=i, column=7).value = bio.ar_mother_lang
        sheet.cell(row=i, column=8).value = bio.ar_other_langs
        sheet.cell(row=i, column=9).value = bio.ar_state
        sheet.cell(row=i, column=10).value = bio.ar_district
        sheet.cell(row=i, column=11).value = bio.surname
        sheet.cell(row=i, column=12).value = str(bio.birthday)
        sheet.cell(row=i, column=13).value = bio.nationality
        sheet.cell(row=i, column=14).value = bio.gender
        sheet.cell(row=i, column=15).value = bio.degree
        sheet.cell(row=i, column=16).value = bio.college
        sheet.cell(row=i, column=17).value = bio.department
        sheet.cell(row=i, column=18).value = bio.mother_lang
        sheet.cell(row=i, column=19).value = bio.other_langs
        sheet.cell(row=i, column=20).value = bio.email
        sheet.cell(row=i, column=21).value = str(bio.create_date)
        sheet.cell(row=i, column=22).value = str(bio.update_date)

    buffer = BytesIO()
    wb.save(buffer)

    buffer.seek(0)

    # create response
    response = HttpResponse(content_type=f'application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="qu-grad-data.xlsx"'
    response.write(buffer.read())

    return response

